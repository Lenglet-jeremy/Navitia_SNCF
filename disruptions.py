from dotenv import load_dotenv
import pandas as pd

import requests
import openpyxl

import json
import os
import re

load_dotenv()

apiKey = os.getenv('APIKEY')
startDate = "20250413T000000"
stopDate = "20250419T000000"
url = f'https://api.sncf.com/v1/coverage/sncf/disruptions//?since={startDate}&until={stopDate}&'

def getDatas():
    datas = []
    headers = {'Authorization': apiKey}
    i = 0
    while True:
        response = requests.get(url + f"start_page={i}", headers=headers)
        response.raise_for_status()

        data = response.json()
        if not data.get('disruptions', []):
            break

        datas.append(data)
        i += 1

    with open('disruptions.json', 'w', encoding='utf-8') as f:
        json.dump(datas, f, ensure_ascii=False, indent=4)

def formatTime(timeStr):
    # Convertit 'hhmmss' vers 'hh:mm:ss'
    if not timeStr or len(timeStr) != 6:
        return timeStr
    return f"{timeStr[:2]}:{timeStr[2:4]}:{timeStr[4:6]}"

def xlsxToJson():

    filePath = "disruptionsDetailed.xlsx"
    if os.path.exists(filePath):
        workbook = openpyxl.load_workbook(filePath)
    else:
        workbook = openpyxl.Workbook()

    if "Disruption_BDD" not in workbook.sheetnames:
        workbook.create_sheet("Disruption_BDD")

    sheet = workbook["Disruption_BDD"]
    rows = []

    # Plage de colonnes à effacer
    minCol = 1
    maxCol = 9

    # Effacer le contenu existant
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=minCol, max_col=maxCol):
        for cell in row:
            cell.value = None

    with open('disruptions.json', 'r', encoding='utf-8') as f:
        rawData = json.load(f)

    for page in rawData:
        for disruption in page.get("disruptions", []):
            disruptionId = disruption.get("disruption_id", "")

            for obj in disruption.get("impacted_objects", []):
                trainId = obj.get("pt_object", {}).get("id", "")

                for stop in obj.get("impacted_stops", []):
                    stopPoint = stop.get("stop_point", {})
                    coord = stopPoint.get("coord", {})

                    # Récupération des temps d'arrivée
                    baseArrivalTime = stop.get("base_arrival_time", "")
                    amendedArrivalTime = stop.get("amended_arrival_time", "")

                    formattedBaseTime = formatTime(baseArrivalTime)
                    formattedAmendedTime = formatTime(amendedArrivalTime)

                    longitude = coord.get('lon', '')
                    latitude = coord.get('lat', '')

                    parts = trainId.split(":")
                    date = next((part for part in parts if re.match(r"\d{4}-\d{2}-\d{2}", part)), "")
                    print(parts)

                    row = {
                        "disruptionId": disruptionId,
                        "cause": stop.get("cause", ""),
                        "trainId": trainId,
                        "stopPointId": stopPoint.get("id", ""),
                        "baseArrivalTime": baseArrivalTime,
                        "amendedArrivalTime": amendedArrivalTime,
                        "formattedBaseArrivalTime": formattedBaseTime,
                        "formattedAmendedArrivalTime": formattedAmendedTime,
                        "lon": longitude,
                        "lat": latitude,
                        "date": date
                    }
                    rows.append(row)

    df = pd.DataFrame(rows)

    # Écriture des en-têtes
    headers = list(df.columns)
    for i, header in enumerate(headers, 1):
        sheet.cell(row=1, column=i).value = header

    # Écriture des données
    for i, rowData in enumerate(df.values, 2):
        for j, value in enumerate(rowData, 1):
            sheet.cell(row=i, column=j).value = value

    workbook.save(filePath)

def main():
    # getDatas()
    xlsxToJson()

if __name__ == "__main__":
    main()
